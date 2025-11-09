"""
Report Export Service - PDF, Excel, CSV Generation
Location: app/services/report_management/report_export_service.py
"""

import io
import csv
from datetime import datetime
from typing import Any, Dict, List

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.enums import ReportTypeEnum


class ReportExportService:
    """Service for exporting reports to various formats"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom PDF styles"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#283593'),
            spaceAfter=12
        )
        
        self.normal_style = self.styles['Normal']
    
    # ==================== PDF Export ==================== #
    
    async def export_to_pdf(
        self,
        data: Any,
        report_type: ReportTypeEnum,
        filters: Dict[str, Any]
    ) -> bytes:
        """Export report data to PDF"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build PDF content
        story = []
        
        # Title
        title = self._get_report_title(report_type)
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 12))
        
        # Report metadata
        metadata = self._build_report_metadata(filters)
        story.append(Paragraph(metadata, self.normal_style))
        story.append(Spacer(1, 20))
        
        # Report content based on type
        if isinstance(data, list):
            # Tabular data
            table = self._build_pdf_table(data, report_type)
            story.append(table)
        elif isinstance(data, dict):
            # Summary/metrics data
            summary = self._build_pdf_summary(data, report_type)
            story.extend(summary)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.normal_style,
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_RIGHT
        )
        story.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(story)
        
        pdf_data = buffer.getvalue()
        buffer.close()
        
        return pdf_data
    
    def _get_report_title(self, report_type: ReportTypeEnum) -> str:
        """Get formatted report title"""
        titles = {
            ReportTypeEnum.daily_sales_summary: "Daily Sales Summary Report",
            ReportTypeEnum.weekly_sales_summary: "Weekly Sales Summary Report",
            ReportTypeEnum.monthly_sales_summary: "Monthly Sales Summary Report",
            ReportTypeEnum.yearly_sales_summary: "Yearly Sales Summary Report",
            ReportTypeEnum.revenue_by_payment_mode: "Revenue by Payment Mode",
            ReportTypeEnum.sales_by_category: "Sales by Medicine Category",
            ReportTypeEnum.top_selling_medicines: "Top Selling Medicines",
            ReportTypeEnum.revenue_trends: "Revenue Trends Analysis",
            ReportTypeEnum.profit_margin_analysis: "Profit Margin Analysis",
            ReportTypeEnum.discount_impact_analysis: "Discount Impact Analysis",
            ReportTypeEnum.coupon_effectiveness: "Coupon Effectiveness Report",
        }
        return titles.get(report_type, report_type.value.replace('_', ' ').title())
    
    def _build_report_metadata(self, filters: Dict[str, Any]) -> str:
        """Build report metadata section"""
        metadata_parts = []
        
        if filters.get('date_from'):
            metadata_parts.append(f"<b>From:</b> {filters['date_from']}")
        if filters.get('date_to'):
            metadata_parts.append(f"<b>To:</b> {filters['date_to']}")
        
        if metadata_parts:
            return " | ".join(metadata_parts)
        return "All Time"
    
    def _build_pdf_table(self, data: List[Dict], report_type: ReportTypeEnum) -> Table:
        """Build PDF table from data"""
        
        if not data:
            return Paragraph("No data available for the selected period.", self.normal_style)
        
        # Get column headers
        headers = list(data[0].keys())
        formatted_headers = [h.replace('_', ' ').title() for h in headers]
        
        # Build table data
        table_data = [formatted_headers]
        
        for row in data:
            formatted_row = []
            for key in headers:
                value = row[key]
                # Format numbers
                if isinstance(value, (int, float)):
                    if 'percentage' in key or 'margin' in key or 'rate' in key:
                        formatted_row.append(f"{value:.2f}%")
                    elif 'revenue' in key or 'amount' in key or 'profit' in key or 'price' in key:
                        formatted_row.append(f"₹{value:,.2f}")
                    else:
                        formatted_row.append(f"{value:,}")
                elif isinstance(value, datetime):
                    formatted_row.append(value.strftime('%Y-%m-%d'))
                else:
                    formatted_row.append(str(value) if value is not None else '-')
            table_data.append(formatted_row)
        
        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f51b5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        return table
    
    def _build_pdf_summary(self, data: Dict, report_type: ReportTypeEnum) -> List:
        """Build PDF summary sections from dict data"""
        
        elements = []
        
        for section_key, section_data in data.items():
            # Section heading
            heading = section_key.replace('_', ' ').title()
            elements.append(Paragraph(heading, self.heading_style))
            elements.append(Spacer(1, 12))
            
            # Section content
            if isinstance(section_data, dict):
                table_data = []
                for key, value in section_data.items():
                    formatted_key = key.replace('_', ' ').title()
                    
                    # Format value
                    if isinstance(value, (int, float)):
                        if 'percentage' in key or 'margin' in key or 'rate' in key:
                            formatted_value = f"{value:.2f}%"
                        elif 'revenue' in key or 'amount' in key or 'profit' in key:
                            formatted_value = f"₹{value:,.2f}"
                        else:
                            formatted_value = f"{value:,}"
                    else:
                        formatted_value = str(value)
                    
                    table_data.append([formatted_key, formatted_value])
                
                if table_data:
                    table = Table(table_data, colWidths=[3*inch, 2*inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ]))
                    elements.append(table)
            
            elements.append(Spacer(1, 20))
        
        return elements
    
    # ==================== Excel Export ==================== #
    
    async def export_to_excel(
        self,
        data: Any,
        report_type: ReportTypeEnum
    ) -> bytes:
        """Export report data to Excel"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report title
        title = self._get_report_title(report_type)
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16, color="1A237E")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Add generation timestamp
        ws.merge_cells('A2:E2')
        timestamp_cell = ws['A2']
        timestamp_cell.value = f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        timestamp_cell.font = Font(italic=True, size=10)
        timestamp_cell.alignment = Alignment(horizontal="center")
        
        # Add data
        start_row = 4
        
        if isinstance(data, list) and data:
            # Tabular data
            headers = list(data[0].keys())
            formatted_headers = [h.replace('_', ' ').title() for h in headers]
            
            # Write headers
            for col_idx, header in enumerate(formatted_headers, 1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows
            for row_idx, row_data in enumerate(data, start_row + 1):
                for col_idx, key in enumerate(headers, 1):
                    value = row_data[key]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Format value
                    if isinstance(value, (int, float)):
                        cell.value = value
                        cell.number_format = '#,##0.00'
                    elif isinstance(value, datetime):
                        cell.value = value
                        cell.number_format = 'yyyy-mm-dd'
                    else:
                        cell.value = value
                    
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15
        
        elif isinstance(data, dict):
            # Summary data
            current_row = start_row
            
            for section_key, section_data in data.items():
                # Section header
                ws.merge_cells(f'A{current_row}:B{current_row}')
                section_cell = ws.cell(row=current_row, column=1)
                section_cell.value = section_key.replace('_', ' ').title()
                section_cell.font = Font(bold=True, size=12, color="283593")
                section_cell.fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
                current_row += 1
                
                # Section data
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        ws.cell(row=current_row, column=1).value = key.replace('_', ' ').title()
                        ws.cell(row=current_row, column=2).value = value
                        
                        if isinstance(value, (int, float)):
                            ws.cell(row=current_row, column=2).number_format = '#,##0.00'
                        
                        current_row += 1
                
                current_row += 1  # Space between sections
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()
        
        return excel_data
    
    # ==================== CSV Export ==================== #
    
    async def export_to_csv(
        self,
        data: Any,
        report_type: ReportTypeEnum
    ) -> bytes:
        """Export report data to CSV"""
        
        buffer = io.StringIO()
        
        if isinstance(data, list) and data:
            # Tabular data
            headers = list(data[0].keys())
            formatted_headers = [h.replace('_', ' ').title() for h in headers]
            
            writer = csv.DictWriter(buffer, fieldnames=headers)
            
            # Write headers
            writer.writerow(dict(zip(headers, formatted_headers)))
            
            # Write data
            for row in data:
                formatted_row = {}
                for key, value in row.items():
                    if isinstance(value, datetime):
                        formatted_row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, (int, float)):
                        formatted_row[key] = f"{value:.2f}"
                    else:
                        formatted_row[key] = value
                writer.writerow(formatted_row)
        
        elif isinstance(data, dict):
            # Summary data - flatten to CSV
            writer = csv.writer(buffer)
            writer.writerow(['Section', 'Metric', 'Value'])
            
            for section_key, section_data in data.items():
                section_name = section_key.replace('_', ' ').title()
                
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        metric_name = key.replace('_', ' ').title()
                        writer.writerow([section_name, metric_name, value])
        
        csv_data = buffer.getvalue().encode('utf-8')
        buffer.close()
        
        return csv_data